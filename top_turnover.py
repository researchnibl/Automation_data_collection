import pandas as pd 
import numpy as np 
from collections import Counter 
import time 
import xlsxwriter
import pandas

from datetime import date
from io import StringIO
import numpy

from openpyxl.utils import get_column_letter
from zbroker import data_collection
from email_sent import email_sent


def broker_name_conversion():
    broker_name_conversion_list = {}
    df = pd.read_csv('Broker_list_sheet.csv')
    for i in range(0, len(df)):
        broker_name_conversion_list[df['B.No'][i]] = df['Broker Name'][i]
    return broker_name_conversion_list

def num_with_comma(x):
  x = int(x*100) / 100
  return "{:,}".format(x)

def sector_conversion1():
    sector_name_conversion = {}
    df = pd.read_csv('sector_list.csv')
    for i in range(0, len(df)):
        sector_name_conversion[df['Stock Symbol '][i].strip()] = df['Sector'][i].strip()
    return sector_name_conversion

def sector_conversion(): 
    with open('symbol.txt') as f:
        contents = f.readlines() 
        f.close()
    Allsectors = {}
    for stock in contents:
        Allsectors[stock.split(",")[0].replace('\n', '').strip()] = stock.split(",")[1].replace('\n', '').strip()
    return Allsectors

#setting broder 
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
broker_name = broker_name_conversion()
total_turnover = data_collection()  

df = pd.read_csv(f'data/buy_sell_broker_details_{date.today()}.csv')

buybroker = df.groupby('Buyer Broker')['Amount'].sum()
buy = buybroker.to_dict()
sellerbroker = df.groupby('Seller Broker')['Amount'].sum()
sell = sellerbroker.to_dict()


df_data = pd.DataFrame()
for i in range(1, 60):
    try:
        df_data.loc[i, 'Broker'] = str(i)
        try:
            df_data.loc[i, 'Broker_Name'] = broker_name[i]
        except:
            df_data.loc[i, 'Broker_Name'] = 'Not Available'
        try:
            
            buy_amount = buy[str(i)]
        except:
            buy_amount = buy[int(i)]
        try:
            
            sell_amount = sell[int(i)]
        except:
            sell_amount = sell[str(i)]
        df_data.loc[i, 'Buy'] = buy_amount
        df_data.loc[i, 'Sell'] = sell_amount
        df_data.loc[i, 'Turnover'] = buy_amount + sell_amount
    except:
        print('NOt available {} broker transactions'.format(i))

df_data.to_csv('all_broker_details.csv')

top_7_broker = df_data.sort_values(by= ['Turnover'], ascending= False)[:7]
top_7_broker_list = top_7_broker['Broker'].to_list()

#TOP-5 Buyer/seller on the basis of Broker
#Top Buyer/seller Top 5 Stocks
all_buyer_details= {}
all_seller_details= {}

for i in range(len(top_7_broker_list)):
  df_top_buy = df[df['Buyer Broker'] == int(top_7_broker_list[i])]
  if df_top_buy.empty:
    df_top_buy = df[df['Buyer Broker'] == str(top_7_broker_list[i])]
    if df_top_buy.empty:
        df_top_buy = df[df['Buyer Broker'] == top_7_broker_list[i]]
        
    try:
        df_top_sell = df[df['Seller Broker'] == int(top_7_broker_list[i])]
        if df_top_sell.empty:
            df_top_sell = df[df['Seller Broker'] == str(top_7_broker_list[i])]
            if df_top_sell.empty:
                df_top_sell = df[df['Seller Broker'] == top_7_broker_list[i]]
    except:
        df_top_sell = df[df['Seller Broker'] == str(' ' + str(top_7_broker_list[(i)])+ ' ')]
    print(df_top_buy)
    print(df_top_sell)
    top_buyer = dict(Counter(dict(df_top_buy.groupby('Stock Symbol')['Amount'].sum())).most_common(5))
    top_seller = dict(Counter(dict(df_top_sell.groupby('Stock Symbol')['Amount'].sum())).most_common(5))
    print(top_buyer, top_seller)
    all_buyer_details['top_{}_buyer'.format(i+1)] = top_buyer
    all_seller_details['top_{}_seller'.format(i+1)] = top_seller
print('All buyer', all_buyer_details)
print('All Seller', all_seller_details)
#top_traded_stock
traded = df.groupby('Stock Symbol')['Amount'].sum()
top_5_traded = dict(Counter(dict(traded)).most_common(5))

def fillXL(worksheet, ColName, From, To, ValuesArr, format1, format2):
  count = 0
  for i in range(From, To+1):
    if len(ValuesArr[count]) != 0:
      if ValuesArr[count] in ['Stock', 'Buy Amount', '% of turnover', 'Sell Amount']:
        worksheet.write(ColName + str(i), ValuesArr[count], format1)
      else:
        worksheet.write(ColName + str(i), ValuesArr[count], format2)
    count += 1

writer = pd.ExcelWriter('Daily_Report/Top 7 broker List with its Turnover ' + time.strftime('%Y-%m-%d', time.localtime(time.time() + 20700)) + '.xlsx', engine='xlsxwriter')
workbook  = writer.book
worksheet = workbook.add_worksheet('Top 7 Broker List')
format = workbook.add_format({'num_format': '0.00'})

worksheet.set_column("A:A", 16)
worksheet.set_column("B:B", 18)
worksheet.set_column(2, 22, 20)

formatting = [
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#F6DDCC', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D5F5E3', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EAECEE', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#FAD7A0', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D6EAF8', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#E8DAEF', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#73C6B6', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EC7063', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
              ]

formattingUnBold = [
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#F6DDCC', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D5F5E3', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EAECEE', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#FAD7A0', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D6EAF8', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#E8DAEF', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#73C6B6', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EC7063', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})
                   ]

formattingpercent = [
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#F6DDCC', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D5F5E3', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EAECEE', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#FAD7A0', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#D6EAF8', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#E8DAEF', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#73C6B6', 'num_format': '0.00%'}),
                    workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#EC7063', 'num_format': '0.00%'})
                   ]


worksheet.write('A1', 'Date: ' + time.strftime('%Y-%m-%d', time.localtime(time.time() + 20700)), formatting[0])
worksheet.write('A2', 'Rank', formatting[0])
worksheet.write('A3', "Broker's Name", formatting[0])
worksheet.write('A4', 'Broker No.', formatting[0])
worksheet.write('A5', 'Total Turnover', formatting[0])
worksheet.write('A6', 'Total Buy', formatting[6])
worksheet.write('A7', 'Total Sell', formatting[7])

worksheet.merge_range('B2:D2', '1', formatting[0])
worksheet.merge_range('E2:G2', '2', formatting[0])
worksheet.merge_range('H2:J2', '3', formatting[0])
worksheet.merge_range('K2:M2', '4', formatting[0])
worksheet.merge_range('N2:P2', '5', formatting[0])
worksheet.merge_range('Q2:S2', '6', formatting[0])
worksheet.merge_range('T2:V2', '7', formatting[0])

worksheet.merge_range('B3:D3', top_7_broker['Broker_Name'].values[0], formatting[1])
worksheet.merge_range('E3:G3', top_7_broker['Broker_Name'].values[1], formatting[2])
worksheet.merge_range('H3:J3', top_7_broker['Broker_Name'].values[2], formatting[3])
worksheet.merge_range('K3:M3', top_7_broker['Broker_Name'].values[3], formatting[4])
worksheet.merge_range('N3:P3', top_7_broker['Broker_Name'].values[4], formatting[5])
worksheet.merge_range('Q3:S3', top_7_broker['Broker_Name'].values[5], formatting[4])
worksheet.merge_range('T3:V3', top_7_broker['Broker_Name'].values[6], formatting[5])


worksheet.merge_range('B4:D4', top_7_broker['Broker'].values[0], formatting[1])
worksheet.merge_range('E4:G4', top_7_broker['Broker'].values[1], formatting[2])
worksheet.merge_range('H4:J4', top_7_broker['Broker'].values[2], formatting[3])
worksheet.merge_range('K4:M4', top_7_broker['Broker'].values[3], formatting[4])
worksheet.merge_range('N4:P4', top_7_broker['Broker'].values[4], formatting[5])
worksheet.merge_range('Q4:S4', top_7_broker['Broker'].values[5], formatting[4])
worksheet.merge_range('T4:V4', top_7_broker['Broker'].values[6], formatting[5])

worksheet.merge_range('B5:D5', num_with_comma(top_7_broker['Turnover'].values[0]), formatting[1])
worksheet.merge_range('E5:G5', num_with_comma(top_7_broker['Turnover'].values[1]), formatting[2])
worksheet.merge_range('H5:J5', num_with_comma(top_7_broker['Turnover'].values[2]), formatting[3])
worksheet.merge_range('K5:M5', num_with_comma(top_7_broker['Turnover'].values[3]), formatting[4])
worksheet.merge_range('N5:P5', num_with_comma(top_7_broker['Turnover'].values[4]), formatting[5])
worksheet.merge_range('Q5:S5', num_with_comma(top_7_broker['Turnover'].values[5]), formatting[4])
worksheet.merge_range('T5:V5', num_with_comma(top_7_broker['Turnover'].values[6]), formatting[5])

worksheet.merge_range('B6:D6', num_with_comma(top_7_broker['Buy'].values[0]), formatting[6])
worksheet.merge_range('E6:G6', num_with_comma(top_7_broker['Buy'].values[1]), formatting[6])
worksheet.merge_range('H6:J6', num_with_comma(top_7_broker['Buy'].values[2]), formatting[6])
worksheet.merge_range('K6:M6', num_with_comma(top_7_broker['Buy'].values[3]), formatting[6])
worksheet.merge_range('N6:P6', num_with_comma(top_7_broker['Buy'].values[4]), formatting[6])
worksheet.merge_range('Q6:S6', num_with_comma(top_7_broker['Buy'].values[5]), formatting[6])
worksheet.merge_range('T6:V6', num_with_comma(top_7_broker['Buy'].values[6]), formatting[6])

worksheet.merge_range('B7:D7', num_with_comma(top_7_broker['Sell'].values[0]), formatting[7])
worksheet.merge_range('E7:G7', num_with_comma(top_7_broker['Sell'].values[1]), formatting[7])
worksheet.merge_range('H7:J7', num_with_comma(top_7_broker['Sell'].values[2]), formatting[7])
worksheet.merge_range('K7:M7', num_with_comma(top_7_broker['Sell'].values[3]), formatting[7])
worksheet.merge_range('N7:P7', num_with_comma(top_7_broker['Sell'].values[4]), formatting[7])
worksheet.merge_range('Q7:S7', num_with_comma(top_7_broker['Sell'].values[5]), formatting[7])
worksheet.merge_range('T7:V7', num_with_comma(top_7_broker['Sell'].values[6]), formatting[7])

worksheet.merge_range('A8:A13', 'Top 5 Buy', formatting[0])
worksheet.merge_range('A15:A20', 'Top 5 Sell', formatting[0])

#add stock Buy amount turnover 
color = 1
for col in range(2, 23, 3):
    if color == 6:
        color = 4 
    char = get_column_letter(col)
    
    worksheet.write(char + '8', 'Stock', formatting[color])
    worksheet.write(get_column_letter(col + 1) + '8', 'Buy Amount', formatting[color])
    worksheet.write(get_column_letter(col + 2) + '8', '% of turnover', formatting[color])
    color = color + 1


#add top_5 Buyer
color = 1
buyer = 1
id = 0
for col in range(2, 23, 3):
    row = 9
    if color == 6:
        color = 4 
    for key, value in all_buyer_details['top_{}_buyer'.format(buyer)].items():
        total_amount = top_7_broker['Turnover'].values[id]
        turnover= value/total_amount 
        worksheet.write(get_column_letter(col) + str(row), key, formattingUnBold[color])
        worksheet.write(get_column_letter(col + 1) + str(row), num_with_comma(value), formattingUnBold[color])
        worksheet.write(get_column_letter(col + 2) + str(row), turnover, formattingpercent[color])
        row = row + 1
    color = color + 1
    id = id + 1
    buyer = buyer + 1

#add stock Sell amount turnover 
color = 1
for col in range(2, 23, 3):
    if color == 6:
        color = 4 
    char = get_column_letter(col)
    worksheet.write(char + '15', 'Stock', formatting[color])
    worksheet.write(get_column_letter(col + 1) + '15', 'Sell Amount', formatting[color])
    worksheet.write(get_column_letter(col + 2) + '15', '% of turnover', formatting[color])
    color += 1

print(all_seller_details)
#add top_5 Seller 
color = 1
seller = 1
id = 0
for col in range(2, 23, 3):
    row = 16
    if color == 6:
        color = 4
    for key, value in all_seller_details['top_{}_seller'.format(seller)].items():
        total_amount = top_7_broker['Turnover'].values[id]
        turnover= value/total_amount
        worksheet.write(get_column_letter(col) + str(row), key, formattingUnBold[color])
        worksheet.write(get_column_letter(col + 1) + str(row), num_with_comma(value), formattingUnBold[color])
        worksheet.write(get_column_letter(col + 2) + str(row), turnover, formattingpercent[color])
        row = row + 1
    color = color + 1
    id = id + 1
    seller = seller + 1


#Turn Over Values added 
bodyTurnFormat = workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})
conclusionTurnFormat = workbook.add_format({'bold': 1, 'border': 1,'align': 'center','valign': 'vcenter', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})
headerTurnFormat = workbook.add_format({'bold': 1,'border': 1,'align': 'center','valign': 'vcenter', 'fg_color': '#FFFF22', 'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})
bodyTurnFormatpercent = workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter', 'num_format': '0.00%'})

worksheet.write('B23', "Top Traded Stocks", headerTurnFormat)
worksheet.write('C23', "Turnover (Rs.)", headerTurnFormat)

worksheet.write('B30', "Sub Indices", headerTurnFormat)
worksheet.write('C30', "Turnover", headerTurnFormat)
worksheet.write('D30', "Percentage (%)", headerTurnFormat)

#top_5_traded companies in a daily market 
count = 24
for key, value in top_5_traded.items():
    worksheet.write('B' + str(count), key, conclusionTurnFormat)
    worksheet.write('C' + str(count), num_with_comma(value), bodyTurnFormat)
    count = count +1

''' Sector Wise Data'''
sector = sector_conversion()

df_sector = df.copy()
for i in range(len(df)):
    try:
        df_sector.loc[i, 'Sector'] = sector[df['Stock Symbol'][i].strip()]
    except:
        df_sector.loc[i, 'Sector'] = 'Promoter Shares/debenture'
        
df_sector.to_csv('sectorwise.csv')
sectorwise = df_sector.groupby('Sector')['Amount'].sum()
sorted_x = sorted(sectorwise.items(), key=lambda kv: kv[1], reverse=True)

c = 31
for i in range(len(sorted_x)):
    perc = sorted_x[i][1] / float(total_turnover.replace(",", ""))
    worksheet.write('B' + str(c), sorted_x[i][0], conclusionTurnFormat)
    worksheet.write('C' + str(c), num_with_comma(sorted_x[i][1]), bodyTurnFormat)
    worksheet.write('D' + str(c), perc, bodyTurnFormatpercent)
    c = c + 1
worksheet.write('B' + str(c), 'Total', conclusionTurnFormat)
worksheet.write('C' + str(c), total_turnover, conclusionTurnFormat)
worksheet.write('D' + str(c), '100%', conclusionTurnFormat)


''' 
TOP 7 Broker List - Sectorwise
All details and new tab with top7 Broker with highest turnover also their highest turnover on the basis of sector. 


'''
worksheet = workbook.add_worksheet('Top 7 Broker List - Sectors')
worksheet.set_column("A:A", 16)
worksheet.set_column(1, 22, 23)
#TOP-5 Buyer/seller on the basis of Broker
#Top Buyer/seller Top 5 Stocks
all_buyer_sector= {}
all_seller_sector= {}
for i in range(len(top_7_broker_list)):
    try:
        df_top_buy = df_sector[df_sector['Buyer Broker'] == str(top_7_broker_list[i])]
        if df_top_buy.empty:
            df_top_buy = df_sector[df_sector['Buyer Broker'] == int(top_7_broker_list[i])]
            if df_top_buy.empty:
                df_top_buy = df_sector[df_sector['Buyer Broker'] == top_7_broker_list[i]]
    except:
        df_top_buy = df_sector[df_sector['Buyer Broker'] == str(' ' + str(top_7_broker_list[i])+ ' ')]
    try:
        df_top_sell = df_sector[df_sector['Seller Broker'] == str(top_7_broker_list[i])]
        if df_top_sell.empty:
            df_top_sell = df_sector[df_sector['Seller Broker'] == int(top_7_broker_list[i])]
            if df_top_sell.empty:
                df_top_sell = df_sector[df_sector['Seller Broker'] == top_7_broker_list[i]]
    except:
        df_top_sell = df_sector[df_sector['Seller Broker'] == str(' ' + str(top_7_broker_list[i])+ ' ')]

    top_buyer = dict(Counter(dict(df_top_buy.groupby('Sector')['Amount'].sum())).most_common(5))
    top_seller = dict(Counter(dict(df_top_sell.groupby('Sector')['Amount'].sum())).most_common(5))
    all_buyer_sector['top_{}_buyer'.format(i+1)] = top_buyer
    all_seller_sector['top_{}_seller'.format(i+1)] = top_seller
    

worksheet.write('A1', 'Date: ' + time.strftime('%Y-%m-%d', time.localtime(time.time() + 20700)), formatting[0])
worksheet.write('A2', 'Rank', formatting[0])
worksheet.write('A3', "Broker's Name", formatting[0])
worksheet.write('A4', 'Broker No.', formatting[0])
worksheet.write('A5', 'Total Turnover', formatting[0])
worksheet.write('A6', 'Total Buy', formatting[6])
worksheet.write('A7', 'Total Sell', formatting[7])

worksheet.merge_range('B2:D2', '1', formatting[0])
worksheet.merge_range('E2:G2', '2', formatting[0])
worksheet.merge_range('H2:J2', '3', formatting[0])
worksheet.merge_range('K2:M2', '4', formatting[0])
worksheet.merge_range('N2:P2', '5', formatting[0])
worksheet.merge_range('Q2:S2', '6', formatting[0])
worksheet.merge_range('T2:V2', '7', formatting[0])

worksheet.merge_range('B3:D3', top_7_broker['Broker_Name'].values[0], formatting[1])
worksheet.merge_range('E3:G3', top_7_broker['Broker_Name'].values[1], formatting[2])
worksheet.merge_range('H3:J3', top_7_broker['Broker_Name'].values[2], formatting[3])
worksheet.merge_range('K3:M3', top_7_broker['Broker_Name'].values[3], formatting[4])
worksheet.merge_range('N3:P3', top_7_broker['Broker_Name'].values[4], formatting[5])
worksheet.merge_range('Q3:S3', top_7_broker['Broker_Name'].values[5], formatting[4])
worksheet.merge_range('T3:V3', top_7_broker['Broker_Name'].values[6], formatting[5])


worksheet.merge_range('B4:D4', top_7_broker['Broker'].values[0], formatting[1])
worksheet.merge_range('E4:G4', top_7_broker['Broker'].values[1], formatting[2])
worksheet.merge_range('H4:J4', top_7_broker['Broker'].values[2], formatting[3])
worksheet.merge_range('K4:M4', top_7_broker['Broker'].values[3], formatting[4])
worksheet.merge_range('N4:P4', top_7_broker['Broker'].values[4], formatting[5])
worksheet.merge_range('Q4:S4', top_7_broker['Broker'].values[5], formatting[4])
worksheet.merge_range('T4:V4', top_7_broker['Broker'].values[6], formatting[5])

worksheet.merge_range('B5:D5', num_with_comma(top_7_broker['Turnover'].values[0]), formatting[1])
worksheet.merge_range('E5:G5', num_with_comma(top_7_broker['Turnover'].values[1]), formatting[2])
worksheet.merge_range('H5:J5', num_with_comma(top_7_broker['Turnover'].values[2]), formatting[3])
worksheet.merge_range('K5:M5', num_with_comma(top_7_broker['Turnover'].values[3]), formatting[4])
worksheet.merge_range('N5:P5', num_with_comma(top_7_broker['Turnover'].values[4]), formatting[5])
worksheet.merge_range('Q5:S5', num_with_comma(top_7_broker['Turnover'].values[5]), formatting[4])
worksheet.merge_range('T5:V5', num_with_comma(top_7_broker['Turnover'].values[6]), formatting[5])

worksheet.merge_range('B6:D6', num_with_comma(top_7_broker['Buy'].values[0]), formatting[6])
worksheet.merge_range('E6:G6', num_with_comma(top_7_broker['Buy'].values[1]), formatting[6])
worksheet.merge_range('H6:J6', num_with_comma(top_7_broker['Buy'].values[2]), formatting[6])
worksheet.merge_range('K6:M6', num_with_comma(top_7_broker['Buy'].values[3]), formatting[6])
worksheet.merge_range('N6:P6', num_with_comma(top_7_broker['Buy'].values[4]), formatting[6])
worksheet.merge_range('Q6:S6', num_with_comma(top_7_broker['Buy'].values[5]), formatting[6])
worksheet.merge_range('T6:V6', num_with_comma(top_7_broker['Buy'].values[6]), formatting[6])

worksheet.merge_range('B7:D7', num_with_comma(top_7_broker['Sell'].values[0]), formatting[7])
worksheet.merge_range('E7:G7', num_with_comma(top_7_broker['Sell'].values[1]), formatting[7])
worksheet.merge_range('H7:J7', num_with_comma(top_7_broker['Sell'].values[2]), formatting[7])
worksheet.merge_range('K7:M7', num_with_comma(top_7_broker['Sell'].values[3]), formatting[7])
worksheet.merge_range('N7:P7', num_with_comma(top_7_broker['Sell'].values[4]), formatting[7])
worksheet.merge_range('Q7:S7', num_with_comma(top_7_broker['Sell'].values[5]), formatting[7])
worksheet.merge_range('T7:V7', num_with_comma(top_7_broker['Sell'].values[6]), formatting[7])

worksheet.merge_range('A8:A13', 'Top 5 Buy', formatting[0])
worksheet.merge_range('A15:A20', 'Top 5 Sell', formatting[0])

#add stock Buy amount turnover 
color = 1
for col in range(2, 23, 3):
    char = get_column_letter(col)
    if color == 6:
        color = 4
    worksheet.write(char + '8', 'Stock', formatting[color])
    worksheet.write(get_column_letter(col + 1) + '8', 'Buy Amount', formatting[color])
    worksheet.write(get_column_letter(col + 2) + '8', '% of turnover', formatting[color])
    color = color + 1


#add top_5 Buyer
color = 1
buyer = 1
id = 0
for col in range(2, 23, 3):
    row = 9
    if color == 6:
        color = 4
    for key, value in all_buyer_sector['top_{}_buyer'.format(buyer)].items():
        total_amount = top_7_broker['Turnover'].values[id]
        turnover= value/total_amount
        worksheet.write(get_column_letter(col) + str(row), key, formattingUnBold[color])
        worksheet.write(get_column_letter(col + 1) + str(row), num_with_comma(value), formattingUnBold[color])
        worksheet.write(get_column_letter(col + 2) + str(row), turnover, formattingpercent[color])
        row = row + 1
    color = color + 1
    id = id + 1
    buyer = buyer + 1

#add stock Sell amount turnover 
count = 1
for col in range(2, 23, 3):
    char = get_column_letter(col)
    if count == 6:
        count = 4
    worksheet.write(char + '15', 'Stock', formatting[count])
    worksheet.write(get_column_letter(col + 1) + '15', 'Sell Amount', formatting[count])
    worksheet.write(get_column_letter(col + 2) + '15', '% of turnover', formatting[count])
    count += 1

print('Buyer', all_buyer_sector)
print('SEller', all_seller_sector)
#add top_5 Seller 
color = 1
seller = 1
id = 0
for col in range(2, 23, 3):
    row = 16
    if color == 6:
        color = 4
    for key, value in all_seller_sector['top_{}_seller'.format(seller)].items():
        total_amount = top_7_broker['Turnover'].values[id]
        turnover= value/total_amount
        worksheet.write(get_column_letter(col) + str(row), key, formattingUnBold[color])
        worksheet.write(get_column_letter(col + 1) + str(row), num_with_comma(value), formattingUnBold[color])
        worksheet.write(get_column_letter(col + 2) + str(row), turnover, formattingpercent[color])
        row = row + 1
    color = color + 1
    id = id + 1
    seller = seller + 1

writer.save()

#Email send Automatic 
# email_sent()
print('Complete all and Sending Email ')
